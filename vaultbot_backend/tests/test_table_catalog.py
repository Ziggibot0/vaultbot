from __future__ import annotations

import pytest
from custom_tools._table_catalog import (
    MAX_SAMPLE_ROWS,
    TableCatalogError,
    inspect_table,
    is_supported_table,
)

pytestmark = pytest.mark.unit


def test_inspect_csv_returns_bounded_schema_and_sample(tmp_path):
    table = tmp_path / "sales.csv"
    table.write_text(
        "Region,Revenue,Returned\nWest,120.5,false\nEast,,true\nWest,80,false\n",
        encoding="utf-8",
    )

    descriptor = inspect_table(table, sample_rows=2)[0]

    assert descriptor.columns == ("Region", "Revenue", "Returned")
    assert descriptor.column_types == ("text", "number", "boolean")
    assert descriptor.row_count == 3
    assert descriptor.null_counts == (0, 1, 0)
    assert len(descriptor.sample_rows) == 2
    assert "Revenue (number)" in descriptor.embedding_text()
    assert "Region=West" in descriptor.embedding_text()


def test_inspect_tsv_normalizes_blank_and_duplicate_headers(tmp_path):
    table = tmp_path / "inventory.tsv"
    table.write_text("SKU\t\tSKU\n10\tWidget\t20\n", encoding="utf-8")

    descriptor = inspect_table(table)[0]

    assert descriptor.columns == ("SKU", "column_2", "SKU_2")
    assert descriptor.column_types == ("integer", "text", "integer")


def test_sample_limit_is_capped(tmp_path):
    table = tmp_path / "large.csv"
    rows = "\n".join(f"{index},value-{index}" for index in range(50))
    table.write_text(f"id,label\n{rows}\n", encoding="utf-8")

    descriptor = inspect_table(table, sample_rows=10_000)[0]

    assert descriptor.row_count == 50
    assert len(descriptor.sample_rows) == MAX_SAMPLE_ROWS


def test_empty_and_unsupported_files_fail_loudly(tmp_path):
    empty = tmp_path / "empty.csv"
    empty.write_text("", encoding="utf-8")
    unsupported = tmp_path / "table.json"
    unsupported.write_text("[]", encoding="utf-8")

    with pytest.raises(TableCatalogError, match="empty"):
        inspect_table(empty)
    with pytest.raises(TableCatalogError, match="Unsupported"):
        inspect_table(unsupported)


def test_supported_extension_check_is_case_insensitive():
    assert is_supported_table("sales.CSV")
    assert is_supported_table("inventory.tsv")
    assert is_supported_table("forecast.XLSX")
    assert not is_supported_table("notes.md")


def test_inspect_xlsx_returns_one_descriptor_per_visible_sheet(tmp_path):
    from openpyxl import Workbook

    table = tmp_path / "sales.xlsx"
    workbook = Workbook()
    quarterly = workbook.active
    quarterly.title = "Quarterly Sales"
    quarterly.append(["Region", "Revenue", "Order Date"])
    quarterly.append(["West", 120.5, "2026-01-15"])
    quarterly.append(["East", None, "2026-01-16"])
    targets = workbook.create_sheet("Targets")
    targets.append(["Region", "Target"])
    targets.append(["West", 100])
    hidden = workbook.create_sheet("Internal")
    hidden.sheet_state = "hidden"
    hidden.append(["Secret"])
    hidden.append(["not indexed"])
    workbook.save(table)

    descriptors = inspect_table(table)

    assert [descriptor.sheet for descriptor in descriptors] == [
        "Quarterly Sales",
        "Targets",
    ]
    assert descriptors[0].columns == ("Region", "Revenue", "Order Date")
    assert descriptors[0].column_types == ("text", "number", "text")
    assert descriptors[0].null_counts == (0, 1, 0)
    assert "Sheet: Quarterly Sales" in descriptors[0].embedding_text()
    assert descriptors[1].row_count == 1
