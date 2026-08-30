"""Bounded metadata extraction for tabular vault files.

The catalog turns a table into compact retrieval text. It never converts the
source to Markdown and never treats an embedded sample as authoritative data;
exact calculations must read the source file through the table query engine.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

SUPPORTED_TABLE_EXTENSIONS = frozenset({".csv", ".tsv", ".xlsx"})
DEFAULT_SAMPLE_ROWS = 5
MAX_SAMPLE_ROWS = 20
MAX_SAMPLE_VALUE_CHARS = 120


class TableCatalogError(ValueError):
    """Raised when a table cannot be inspected safely."""


@dataclass(frozen=True)
class TableDescriptor:
    """Compact, serializable description of one logical table."""

    file_path: str
    name: str
    format: str
    sheet: str | None
    columns: tuple[str, ...]
    column_types: tuple[str, ...]
    row_count: int
    null_counts: tuple[int, ...]
    sample_rows: tuple[tuple[str, ...], ...]

    def to_metadata(self) -> dict[str, Any]:
        return {
            "source_type": "table",
            "file_path": self.file_path,
            "name": self.name,
            "format": self.format,
            "sheet": self.sheet,
            "columns": list(self.columns),
            "column_types": list(self.column_types),
            "row_count": self.row_count,
            "null_counts": dict(zip(self.columns, self.null_counts, strict=True)),
            "sample_rows": [list(row) for row in self.sample_rows],
        }

    def embedding_text(self) -> str:
        schema = ", ".join(
            f"{name} ({kind})"
            for name, kind in zip(self.columns, self.column_types, strict=True)
        )
        lines = [
            f"Table: {self.name}",
            f"Format: {self.format}",
            f"Rows: {self.row_count}",
            f"Columns: {schema}",
        ]
        if self.sheet:
            lines.insert(2, f"Sheet: {self.sheet}")
        if self.sample_rows:
            lines.append("Representative values:")
            for row in self.sample_rows:
                values = ", ".join(
                    f"{column}={value}"
                    for column, value in zip(self.columns, row, strict=True)
                    if value
                )
                if values:
                    lines.append(values)
        return "\n".join(lines)


def is_supported_table(path: str | Path) -> bool:
    return Path(path).suffix.lower() in SUPPORTED_TABLE_EXTENSIONS


def inspect_table(
    path: str | Path, *, sample_rows: int = DEFAULT_SAMPLE_ROWS
) -> list[TableDescriptor]:
    """Inspect a supported file and return its logical tables.

    CSV and TSV contain one logical table. XLSX returns one descriptor per
    visible, non-empty worksheet.
    """
    source = Path(path).resolve()
    if not source.is_file():
        raise TableCatalogError(f"Table file does not exist: {source}")
    suffix = source.suffix.lower()
    if suffix not in SUPPORTED_TABLE_EXTENSIONS:
        raise TableCatalogError(f"Unsupported table format: {suffix or '(none)'}")
    bounded_sample_rows = max(0, min(int(sample_rows), MAX_SAMPLE_ROWS))
    if suffix == ".xlsx":
        return _inspect_xlsx(source, bounded_sample_rows)
    delimiter = "\t" if suffix == ".tsv" else ","
    return [_inspect_delimited(source, delimiter, bounded_sample_rows)]


def _inspect_delimited(
    source: Path, delimiter: str, sample_limit: int
) -> TableDescriptor:
    try:
        handle = source.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise TableCatalogError(f"Could not open table: {exc}") from exc

    with handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise TableCatalogError("Table is empty") from exc
        columns = _normalize_headers(raw_headers)
        if not columns:
            raise TableCatalogError("Table has no columns")

        observed: list[list[str]] = [[] for _ in columns]
        null_counts = [0 for _ in columns]
        samples: list[tuple[str, ...]] = []
        row_count = 0
        for raw_row in reader:
            if not raw_row or not any(value.strip() for value in raw_row):
                continue
            row_count += 1
            row = [
                raw_row[index].strip() if index < len(raw_row) else ""
                for index in range(len(columns))
            ]
            for index, value in enumerate(row):
                if value:
                    if len(observed[index]) < 100:
                        observed[index].append(value)
                else:
                    null_counts[index] += 1
            if len(samples) < sample_limit:
                samples.append(tuple(value[:MAX_SAMPLE_VALUE_CHARS] for value in row))

    return TableDescriptor(
        file_path=str(source),
        name=source.stem,
        format=source.suffix.lower().lstrip("."),
        sheet=None,
        columns=columns,
        column_types=tuple(_infer_type(values) for values in observed),
        row_count=row_count,
        null_counts=tuple(null_counts),
        sample_rows=tuple(samples),
    )


def _inspect_xlsx(source: Path, sample_limit: int) -> list[TableDescriptor]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - decoder errors become catalog errors
        raise TableCatalogError(f"Could not open workbook: {exc}") from exc

    descriptors: list[TableDescriptor] = []
    try:
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            rows = worksheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration:
                continue
            columns = _normalize_headers([_cell_text(value) for value in raw_headers])
            if not columns or not any(raw_headers):
                continue

            observed: list[list[str]] = [[] for _ in columns]
            null_counts = [0 for _ in columns]
            samples: list[tuple[str, ...]] = []
            row_count = 0
            for raw_row in rows:
                row = [
                    _cell_text(raw_row[index]) if index < len(raw_row) else ""
                    for index in range(len(columns))
                ]
                if not any(row):
                    continue
                row_count += 1
                for index, value in enumerate(row):
                    if value:
                        if len(observed[index]) < 100:
                            observed[index].append(value)
                    else:
                        null_counts[index] += 1
                if len(samples) < sample_limit:
                    samples.append(
                        tuple(value[:MAX_SAMPLE_VALUE_CHARS] for value in row)
                    )

            descriptors.append(
                TableDescriptor(
                    file_path=str(source),
                    name=f"{source.stem} - {worksheet.title}",
                    format="xlsx",
                    sheet=worksheet.title,
                    columns=columns,
                    column_types=tuple(_infer_type(values) for values in observed),
                    row_count=row_count,
                    null_counts=tuple(null_counts),
                    sample_rows=tuple(samples),
                )
            )
    finally:
        workbook.close()

    if not descriptors:
        raise TableCatalogError("Workbook has no visible, non-empty sheets")
    return descriptors


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _normalize_headers(headers: list[str]) -> tuple[str, ...]:
    normalized: list[str] = []
    counts: dict[str, int] = {}
    for index, raw_header in enumerate(headers, start=1):
        base = raw_header.strip() or f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        normalized.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return tuple(normalized)


def _infer_type(values: list[str]) -> str:
    if not values:
        return "unknown"
    if all(_is_integer(value) for value in values):
        return "integer"
    if all(_is_number(value) for value in values):
        return "number"
    if all(value.casefold() in {"true", "false"} for value in values):
        return "boolean"
    return "text"


def _is_integer(value: str) -> bool:
    try:
        int(value)
    except ValueError:
        return False
    return True


def _is_number(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True
