"""Constrained, read-only queries over tabular vault files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import duckdb
from custom_tools._table_catalog import (
    TableCatalogError,
    inspect_table,
    is_supported_table,
)

MAX_RESULT_ROWS = 100
AGGREGATIONS = frozenset({"sum", "avg", "min", "max", "count"})
FILTER_OPERATORS = frozenset(
    {"eq", "ne", "gt", "gte", "lt", "lte", "contains", "is_null", "not_null"}
)
JOIN_TYPES = frozenset({"inner", "left"})


class TableQueryError(ValueError):
    """Raised when a table query is invalid or unsafe."""


def resolve_vault_table(vault_root: str | Path, source: str) -> Path:
    root = Path(vault_root).resolve()
    candidate = (root / source).resolve()
    try:
        candidate.relative_to(root)
    except ValueError as exc:
        raise TableQueryError("Table source must be inside the vault") from exc
    if not candidate.is_file():
        raise TableQueryError(f"Table source does not exist: {source}")
    if not is_supported_table(candidate):
        raise TableQueryError(f"Unsupported table format: {candidate.suffix}")
    return candidate


def query_table(
    vault_root: str | Path,
    *,
    source: str,
    operation: str,
    sheet: str | None = None,
    columns: list[str] | None = None,
    aggregate: str | None = None,
    group_by: list[str] | None = None,
    filters: list[dict[str, Any]] | None = None,
    join_source: str | None = None,
    join_sheet: str | None = None,
    left_on: str | None = None,
    right_on: str | None = None,
    right_columns: list[str] | None = None,
    join_type: str = "inner",
    limit: int = 20,
) -> dict[str, Any]:
    path = resolve_vault_table(vault_root, source)
    descriptors = inspect_table(path)
    descriptor = _select_descriptor(descriptors, sheet)
    if operation == "inspect":
        return {
            "status": "success",
            "operation": operation,
            "table": descriptor.to_metadata(),
            "provenance": _provenance(path, descriptor, operation),
        }

    bounded_limit = max(1, min(int(limit), MAX_RESULT_ROWS))
    connection = duckdb.connect(":memory:")
    joined_path: Path | None = None
    joined_descriptor = None
    try:
        _load_source(connection, path, descriptor.sheet, "source_table")
        available = set(descriptor.columns)
        requested = columns or list(descriptor.columns)
        _validate_columns(requested, available)
        where_sql, parameters = _build_filters(
            filters or [], available, table_alias="l" if operation == "join" else None
        )
        if operation == "preview":
            sql = (
                f"SELECT {', '.join(_quote(column) for column in requested)} "
                f"FROM source_table{where_sql} LIMIT {bounded_limit}"
            )
        elif operation == "aggregate":
            if aggregate not in AGGREGATIONS:
                raise TableQueryError(
                    f"aggregate must be one of {sorted(AGGREGATIONS)}"
                )
            if len(requested) != 1:
                raise TableQueryError("aggregate requires exactly one value column")
            groups = group_by or []
            _validate_columns(groups, available)
            select_groups = [_quote(column) for column in groups]
            value = _quote(requested[0])
            expression = f"{aggregate.upper()}({value}) AS value"
            select = ", ".join([*select_groups, expression])
            sql = f"SELECT {select} FROM source_table{where_sql}"
            if select_groups:
                sql += f" GROUP BY {', '.join(select_groups)}"
            sql += f" LIMIT {bounded_limit}"
        elif operation == "join":
            if not join_source or not left_on or not right_on:
                raise TableQueryError(
                    "join requires join_source, left_on, and right_on"
                )
            if join_type not in JOIN_TYPES:
                raise TableQueryError(f"join_type must be one of {sorted(JOIN_TYPES)}")
            joined_path = resolve_vault_table(vault_root, join_source)
            joined_descriptor = _select_descriptor(
                inspect_table(joined_path), join_sheet
            )
            right_available = set(joined_descriptor.columns)
            _validate_columns([left_on], available)
            _validate_columns([right_on], right_available)
            selected_right = right_columns or list(joined_descriptor.columns)
            _validate_columns(selected_right, right_available)
            _load_source(
                connection,
                joined_path,
                joined_descriptor.sheet,
                "right_table",
            )
            left_select = [
                f"l.{_quote(column)} AS {_quote('left__' + column)}"
                for column in requested
            ]
            right_select = [
                f"r.{_quote(column)} AS {_quote('right__' + column)}"
                for column in selected_right
            ]
            selected = ", ".join([*left_select, *right_select])
            sql = (
                f"SELECT {selected} FROM source_table AS l "
                f"{join_type.upper()} JOIN right_table AS r "
                f"ON l.{_quote(left_on)} = r.{_quote(right_on)}"
                f"{where_sql} LIMIT {bounded_limit}"
            )
        else:
            raise TableQueryError(
                "operation must be inspect, preview, aggregate, or join"
            )

        cursor = connection.execute(sql, parameters)
        result_columns = [item[0] for item in cursor.description]
        rows = [list(row) for row in cursor.fetchall()]
    except duckdb.Error as exc:
        raise TableQueryError(f"Table query failed: {exc}") from exc
    finally:
        connection.close()

    provenance = _provenance(path, descriptor, operation)
    provenance.update(
        {
            "columns": requested,
            "group_by": group_by or [],
            "aggregate": aggregate,
            "filters": filters or [],
            "input_rows": descriptor.row_count,
            "output_rows": len(rows),
        }
    )
    if joined_path is not None and joined_descriptor is not None:
        provenance["join"] = {
            "source": str(joined_path),
            "sheet": joined_descriptor.sheet,
            "left_on": left_on,
            "right_on": right_on,
            "join_type": join_type,
            "input_rows": joined_descriptor.row_count,
        }
    return {
        "status": "success",
        "operation": operation,
        "columns": result_columns,
        "rows": rows,
        "provenance": provenance,
    }


def _select_descriptor(descriptors, sheet: str | None):
    if sheet:
        for descriptor in descriptors:
            if descriptor.sheet == sheet:
                return descriptor
        raise TableQueryError(f"Workbook has no sheet named: {sheet}")
    if len(descriptors) > 1:
        names = [descriptor.sheet for descriptor in descriptors]
        raise TableQueryError(f"Workbook has multiple sheets; choose one of {names}")
    return descriptors[0]


def _load_source(connection, path: Path, sheet: str | None, table_name: str) -> None:
    if path.suffix.lower() in {".csv", ".tsv"}:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        connection.read_csv(str(path), delimiter=delimiter).create_view(table_name)
        return

    workbook = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet else workbook.active
        if worksheet is None:
            raise TableQueryError("Workbook has no active sheet")
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows)]
        normalized = inspect_table(path)
        descriptor = next(item for item in normalized if item.sheet == worksheet.title)
        duckdb_types = {
            "integer": "BIGINT",
            "number": "DOUBLE",
            "boolean": "BOOLEAN",
            "text": "VARCHAR",
            "unknown": "VARCHAR",
        }
        column_defs = ", ".join(
            f"{_quote(name)} {duckdb_types[column_type]}"
            for name, column_type in zip(
                descriptor.columns, descriptor.column_types, strict=True
            )
        )
        connection.execute(f"CREATE TABLE {_quote(table_name)} ({column_defs})")
        placeholders = ", ".join("?" for _ in headers)
        values = [
            tuple(
                row[index] if index < len(row) else None
                for index in range(len(headers))
            )
            for row in rows
            if any(value is not None for value in row)
        ]
        if values:
            connection.executemany(
                f"INSERT INTO {_quote(table_name)} VALUES ({placeholders})", values
            )
    except (KeyError, StopIteration, TableCatalogError) as exc:
        raise TableQueryError(f"Could not load workbook sheet: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _validate_columns(columns: list[str], available: set[str]) -> None:
    unknown = [column for column in columns if column not in available]
    if unknown:
        raise TableQueryError(f"Unknown columns: {unknown}")


def _build_filters(
    filters: list[dict[str, Any]],
    available: set[str],
    table_alias: str | None = None,
) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    parameters: list[Any] = []
    comparisons = {
        "eq": "=",
        "ne": "!=",
        "gt": ">",
        "gte": ">=",
        "lt": "<",
        "lte": "<=",
    }
    for item in filters:
        if not isinstance(item, dict):
            raise TableQueryError("Each filter must be an object")
        column = item.get("column")
        operator = item.get("operator")
        if column not in available:
            raise TableQueryError(f"Unknown filter column: {column}")
        if operator not in FILTER_OPERATORS:
            raise TableQueryError(
                f"filter operator must be one of {sorted(FILTER_OPERATORS)}"
            )
        identifier = _quote(column)
        if table_alias:
            identifier = f"{table_alias}.{identifier}"
        if operator == "is_null":
            clauses.append(f"{identifier} IS NULL")
        elif operator == "not_null":
            clauses.append(f"{identifier} IS NOT NULL")
        elif operator == "contains":
            clauses.append(f"contains(CAST({identifier} AS VARCHAR), ?)")
            parameters.append(str(item.get("value", "")))
        else:
            if "value" not in item:
                raise TableQueryError(f"Filter operator {operator} requires a value")
            clauses.append(f"{identifier} {comparisons[operator]} ?")
            parameters.append(item["value"])
    if not clauses:
        return "", parameters
    return " WHERE " + " AND ".join(clauses), parameters


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _provenance(path, descriptor, operation: str) -> dict[str, Any]:
    return {
        "source": str(path),
        "sheet": descriptor.sheet,
        "operation": operation,
    }
